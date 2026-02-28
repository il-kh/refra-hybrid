from pathlib import Path
from openpyxl import Workbook

from analysis_itm import WingResult
from analysis_plusminus import PMPairResult

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ===========================================================================
# Excel writer
# ===========================================================================

_HEADERS = [
    "Transect", "File", "Shot pos (m)",
    "V1 right (m/s)", "V2 right (m/s)", "t_i right (ms)", "Depth right (m)",
    "V2 pts R", "xc cov R (%)",
    "V1 left (m/s)",  "V2 left (m/s)",  "t_i left (ms)",  "Depth left (m)",
    "V2 pts L", "xc cov L (%)",
    "Depth avg (m)", "Warnings",
]
_COL_WIDTHS = [10, 24, 13, 14, 14, 14, 14, 9, 12, 13, 13, 13, 13, 9, 12, 13, 50]


def _wing_val(wing: WingResult, key: str, decimals: int = 2):
    """Safe getter – returns 'N/A' for missing / NaN values."""
    if wing is None:
        return 'N/A'
    val = wing.get(key, float('nan'))
    if isinstance(val, float):
        return round(val, decimals) if not np.isnan(val) else 'N/A'
    return val


def _wing_pct(wing: WingResult, key: str):
    """Format a ratio as a percentage integer, or 'N/A'."""
    if wing is None:
        return 'N/A'
    val = wing.get(key, 0.0)
    if val == 0 or np.isnan(val):
        return 'N/A'
    return round(float(val) * 100)


def save_excel(results: list[dict], xlsx_path: Path) -> None:
    """Write a formatted summary spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Refraction ITM Results"

    thin      = Side(style='thin')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre    = Alignment(horizontal="center")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")

    for col, (header, width) in enumerate(
            zip(_HEADERS, _COL_WIDTHS), start=1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30

    for row_num, r in enumerate(results, start=2):
        res   = r['res']
        right = res.get('right')
        left  = res.get('left')
        depths = [w['depth'] for w in filter(None, (right, left))]
        d_avg  = float(np.nanmean(depths)) if depths else float('nan')

        row_values = [
            r['transect_id'], r['file_name'],
            round(res['true_shot_loc'], 3),
            _wing_val(right, 'v1', 1), _wing_val(right, 'v2', 1),
            _wing_val(right, 't_i_ms', 3), _wing_val(right, 'depth', 3),
            _wing_val(right, 'v2_count', 0), _wing_pct(right, 'xc_ratio'),
            _wing_val(left,  'v1', 1), _wing_val(left,  'v2', 1),
            _wing_val(left,  't_i_ms', 3), _wing_val(left,  'depth', 3),
            _wing_val(left, 'v2_count', 0), _wing_pct(left, 'xc_ratio'),
            round(d_avg, 3) if not np.isnan(d_avg) else 'N/A',
            ' | '.join(res.get('warnings', [])),
        ]
        has_warnings = bool(res.get('warnings'))
        for col, value in enumerate(row_values, start=1):
            cell           = ws.cell(row=row_num, column=col, value=value)
            cell.border    = border
            cell.alignment = centre
            if has_warnings:
                cell.fill = warn_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"
    wb.save(xlsx_path)
    print(f"  Excel saved → {xlsx_path}")


# ===========================================================================
# Plus-Minus Excel writer
# ===========================================================================

_PM_HEADERS = [
    "Transect", "Shot A", "Shot B",
    "Shot A pos (m)", "Shot B pos (m)", "T_AB (ms)",
    "V2 (m/s)", "V2 R²", "V1 used (m/s)",
    "Geo X (m)", "Depth (m)", "z_rock (m ASL)",
    "Warnings",
]
_PM_COL_WIDTHS = [10, 22, 22, 14, 14, 12, 12, 10, 12, 12, 12, 14, 50]


def save_pm_excel(
        pm_results: dict[int, list[PMPairResult]],
        elev_data: dict[int, tuple[np.ndarray, np.ndarray]],
        xlsx_path: Path,
) -> None:
    """Write Plus-Minus results to a formatted Excel workbook."""
    from elevation import interpolate_elevation

    wb = Workbook()
    ws = wb.active
    ws.title = "Plus-Minus Results"

    thin   = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")

    for col, (header, width) in enumerate(
            zip(_PM_HEADERS, _PM_COL_WIDTHS), start=1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="5B2C6F")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30

    row_num = 2
    for tid in sorted(pm_results):
        elev_x, elev_z = elev_data.get(tid, (np.array([]), np.array([])))
        for pr in pm_results[tid]:
            has_warn = bool(pr.warnings)
            warn_str = ' | '.join(pr.warnings) if pr.warnings else ''

            for i, (x, d) in enumerate(zip(pr.geo_x, pr.depths)):
                z_rock = ''
                if not np.isnan(d) and len(elev_x) > 0:
                    z_s = interpolate_elevation(elev_x, elev_z, float(x))
                    if z_s is not None:
                        z_rock = round(z_s - d, 3)

                row_values = [
                    tid,
                    pr.file_a if i == 0 else '',
                    pr.file_b if i == 0 else '',
                    round(pr.shot_a, 2) if i == 0 else '',
                    round(pr.shot_b, 2) if i == 0 else '',
                    round(pr.t_ab_ms, 2) if i == 0 else '',
                    round(pr.v2, 1) if i == 0 and not np.isnan(pr.v2) else ('' if i else 'N/A'),
                    round(pr.v2_r2, 4) if i == 0 and not np.isnan(pr.v2_r2) else '',
                    round(pr.v1_used, 1) if i == 0 else '',
                    round(float(x), 2),
                    round(float(d), 3) if not np.isnan(d) else 'N/A',
                    z_rock,
                    warn_str if i == 0 else '',
                ]
                for col, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border    = border
                    cell.alignment = centre
                    if has_warn:
                        cell.fill = warn_fill
                row_num += 1

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"
    wb.save(xlsx_path)
    print(f"  PM Excel saved → {xlsx_path}")
