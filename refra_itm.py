"""
refra_itm.py
============
Seismic refraction analysis using the Intercept-Time Method (ITM).

Workflow
--------
1. Read first-arrival picks from .vs files grouped by transect.
2. For each shot, split the spread into left / right wings and fit a
   two-segment linear travel-time curve (V1 = direct wave, V2 = refracted).
3. Derive depth-to-refractor from the intercept-time formula.
4. Export results to:
   - PDF  : travel-time plots  (refra_itm_traveltimes.pdf)
   - PDF  : elevation + rock-depth profiles  (refra_itm_elevation.pdf)
   - XLSX : tabular summary  (refra_itm_results.xlsx)
"""

import csv
import sys
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as pdf_backend
from scipy.stats import linregress
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
V1_MIN: float = 150.0
V1_MAX: float = 800.0
V2_MIN: float = 800.0
V2_MAX: float = 3000.0
V2_ROCK_MIN: float = 800.0

BP_SLOPE_TOLERANCE: float = 1.20
MIN_SEGMENT_POINTS: int = 3

# Elevation plot axis limits – 1:1 aspect is enforced by set_aspect('equal').
# x: 0 → 28 m  (28 m range)
# y: -15 → +7 m (22 m range)  — clipped at -15 for readability
ELEV_X_MIN: float =  0.0
ELEV_X_MAX: float = 28.0
ELEV_Y_MIN: float = -15.0
ELEV_Y_MAX: float =  7.0

SHEET_PILE_DEPTH: float = 18.0  # m below surface

# Geotech style
_DPL_COLOR:        str = 'darkorange'
_CPTU_COLOR:       str = 'royalblue'
_ROCK_MARKER_SIZE: int = 7
_SHEET_PILE_COLOR: str = 'black'

# ---------------------------------------------------------------------------
# Type aliases
# ---------------------------------------------------------------------------
WingResult = Optional[dict]
ShotResult = dict


# ===========================================================================
# Elevation data  (read from elev.csv)
# ===========================================================================

def read_elev_csv(csv_path: Path) -> dict[int, tuple[np.ndarray, np.ndarray]]:
    """
    Parse elev.csv into per-transect (distances, elevations) arrays.

    Relevant columns
    ----------------
    line_no         : transect id  (e.g. '0220' or '220')
    dist_first_gp_m : raw distance along transect (m)
    geop_id         : integer geophone id; the row where geop_id == 0
                      defines x = 0 on the plot axis.  All other x values
                      are shifted by the same offset so that the geop_id==0
                      row lands exactly at x = 0.
    z               : elevation (m ASL)

    Rows with x < ELEV_X_MIN or x > ELEV_X_MAX are kept so that
    np.interp can extrapolate (clamp) the surface elevation at the
    plot boundaries when needed.

    Returns
    -------
    dict  transect_id → (dist_array, elev_array)  sorted by x
    """
    # raw storage: tid → list of (raw_dist, geop_id_or_None, z)
    raw: dict[int, list[tuple[float, Optional[float], float]]] = {}

    if not csv_path.exists():
        print(f"  ⚠  elev.csv not found at {csv_path} – no elevation data.")
        return {}

    with open(csv_path, newline='', encoding='utf-8-sig') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            raw_line = row.get('line_no', '').strip()
            if not raw_line:
                continue
            try:
                tid = int(raw_line)
            except ValueError:
                continue

            dist = _parse_float(row.get('dist_first_gp_m', ''))
            z    = _parse_float(row.get('z', ''))
            if dist is None or z is None:
                continue

            geop_id = _parse_float(row.get('geop_id', ''))
            raw.setdefault(tid, []).append((dist, geop_id, z))

    result: dict[int, tuple[np.ndarray, np.ndarray]] = {}

    for tid, pts in raw.items():
        # Find the raw dist value for the row where geop_id == 0
        ref_dist: Optional[float] = None
        for dist, geop_id, _ in pts:
            if geop_id is not None and geop_id == 0.0:
                ref_dist = dist
                break

        if ref_dist is None:
            print(f"  ⚠  Transect {tid}: no geop_id==0 row found in "
                  f"elev.csv – using raw distances unchanged.")
            ref_dist = 0.0

        # Apply shift so that geop_id==0 lands at x=0
        shifted = sorted(
            [(dist - ref_dist, z) for dist, _, z in pts],
            key=lambda p: p[0],
        )

        result[tid] = (
            np.array([p[0] for p in shifted]),
            np.array([p[1] for p in shifted]),
        )

        x_min_data = result[tid][0].min()
        x_max_data = result[tid][0].max()
        print(f"  Transect {tid}: {len(shifted)} elevation points, "
              f"x = {x_min_data:.1f} … {x_max_data:.1f} m "
              f"(ref shift = {-ref_dist:+.2f} m)")

    print(f"  elev.csv: elevation data loaded for "
          f"{len(result)} transect(s).")
    return result


# ===========================================================================
# Geotech data
# ===========================================================================

@dataclass
class GeotechPoint:
    """One row from geotech.csv that carries an actual test result."""
    test_type:     str
    line_no:       int
    dist_x:        float           # dist_first_gp_m → x-axis (m)
    tested_depth:  float           # m below surface
    depth_of_rock: Optional[float] # m below surface, or None


@dataclass
class SheetPileLine:
    """Position of the planned sheet-pile wall for one transect."""
    line_no: int
    dist_x:  float  # dist_first_gp_m where dist_sheet_pile_m == 0


def _parse_float(value: str) -> Optional[float]:
    """Return float or None for blank / non-numeric CSV cells."""
    v = value.strip()
    if not v:
        return None
    try:
        return float(v)
    except ValueError:
        return None


def read_geotech_csv(csv_path: Path
                     ) -> tuple[dict[int, list[GeotechPoint]],
                                dict[int, SheetPileLine]]:
    """
    Parse geotech.csv.

    Returns
    -------
    geotech_by_tid   : transect_id → list[GeotechPoint]
    sheetpile_by_tid : transect_id → SheetPileLine
    """
    geotech:   dict[int, list[GeotechPoint]] = {}
    sheetpile: dict[int, SheetPileLine]      = {}

    if not csv_path.exists():
        print(f"  ⚠  geotech.csv not found at {csv_path} – skipping.")
        return geotech, sheetpile

    with open(csv_path, newline='', encoding='utf-8-sig') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            raw_line = row.get('line_no', '').strip()
            if not raw_line:
                continue
            try:
                tid = int(raw_line)
            except ValueError:
                continue

            dist_sheet = _parse_float(row.get('dist_sheet_pile_m', ''))
            dist_x     = _parse_float(row.get('dist_first_gp_m', ''))
            if dist_x is None:
                continue

            # Sheet-pile reference (dist_sheet_pile_m == 0)
            if dist_sheet is not None and dist_sheet == 0.0:
                if tid not in sheetpile:
                    sheetpile[tid] = SheetPileLine(line_no=tid, dist_x=dist_x)

            # Actual test row
            test_type = row.get('test_type', '').strip()
            if not test_type:
                continue

            tested = _parse_float(row.get('tested_depth_m', ''))
            rock   = _parse_float(row.get('depth_of_rock_m', ''))
            if tested is None:
                continue

            geotech.setdefault(tid, []).append(GeotechPoint(
                test_type     = test_type,
                line_no       = tid,
                dist_x        = dist_x,
                tested_depth  = tested,
                depth_of_rock = rock,
            ))

    print(f"  Geotech CSV: {sum(len(v) for v in geotech.values())} tests "
          f"across {len(geotech)} transect(s); "
          f"{len(sheetpile)} sheet-pile position(s) loaded.")
    return geotech, sheetpile


# ===========================================================================
# I/O helpers
# ===========================================================================

def read_vs_file(file_path: Path) -> tuple[Optional[float],
                                           np.ndarray, np.ndarray]:
    """
    Parse one .vs pick file.

    Lines
    -----
    1-2  : header (ignored)
    3    : shot position in column 0
    4-27 : geophone_distance  travel_time_ms  [weight]
    """
    shot_pos:  Optional[float] = None
    distances: list[float] = []
    times_ms:  list[float] = []

    with open(file_path, 'r') as fh:
        for line_num, line in enumerate(fh, start=1):
            row = line.strip().split()
            if line_num < 3:
                continue
            if line_num == 3:
                shot_pos = float(row[0])
                continue
            if line_num > 27:
                break
            if len(row) >= 2:
                try:
                    distances.append(float(row[0]))
                    times_ms.append(float(row[1]))
                except ValueError:
                    print(f"  Warning: non-numeric data on line {line_num} "
                          f"of {file_path.name} – skipped.")

    return shot_pos, np.array(distances), np.array(times_ms)


def group_by_transect(vs_files: list[Path]) -> dict[int, list[Path]]:
    """Group .vs files by transect id from filenames ``<tid>-<sid>.vs``."""
    pattern = re.compile(r'^(\d+)-(\d+)\.vs$', re.IGNORECASE)
    groups: dict[int, list[Path]] = {}
    for path in vs_files:
        m = pattern.match(path.name)
        if not m:
            print(f"  Skipping unrecognised filename: {path.name}")
            continue
        groups.setdefault(int(m.group(1)), []).append(path)
    return dict(sorted(groups.items()))


# ===========================================================================
# Elevation helpers
# ===========================================================================

def interpolate_elevation(elev_x: np.ndarray, elev_z: np.ndarray,
                          query_x: float) -> Optional[float]:
    """Linearly interpolate surface elevation at *query_x*."""
    if len(elev_x) == 0:
        return None
    return float(np.interp(query_x, elev_x, elev_z))


# ===========================================================================
# Geometry helpers
# ===========================================================================

def compute_offsets(geophone_locs: np.ndarray,
                    times_ms: np.ndarray,
                    shot_pos: Optional[float] = None
                    ) -> tuple[np.ndarray, float, int]:
    """
    Compute absolute source–receiver offsets.

    Returns (offsets, true_shot_loc, shot_idx).
    """
    shot_idx    = int(np.argmin(times_ms))
    nearest_geo = geophone_locs[shot_idx]
    true_shot   = (float(shot_pos)
                   if shot_pos is not None and shot_pos != 0.0
                   else nearest_geo)
    return np.abs(geophone_locs - true_shot), true_shot, shot_idx


# ===========================================================================
# Breakpoint detection
# ===========================================================================

def _ssr_breakpoint(wing_offsets: np.ndarray,
                    wing_times_sec: np.ndarray,
                    min_points: int) -> int:
    """Fallback: minimum-SSR two-segment split."""
    n = len(wing_offsets)
    best_ssr, best_idx = np.inf, n - 2

    for idx in range(min_points, n - min_points + 1):
        v2_off, v2_t = wing_offsets[idx:], wing_times_sec[idx:]
        if len(v2_off) < 2:
            continue
        sl, *_ = linregress(v2_off, v2_t)
        if sl <= 0:
            continue
        v1_fit = np.polyval(
            np.polyfit(wing_offsets[:idx], wing_times_sec[:idx], 1),
            wing_offsets[:idx])
        v2_fit = np.polyval(np.polyfit(v2_off, v2_t, 1), v2_off)
        ssr = (np.sum((wing_times_sec[:idx] - v1_fit) ** 2) +
               np.sum((v2_t - v2_fit) ** 2))
        if ssr < best_ssr:
            best_ssr, best_idx = ssr, idx
    return best_idx


def find_breakpoint_on_wing(wing_offsets: np.ndarray,
                             wing_times_sec: np.ndarray,
                             min_points: int = MIN_SEGMENT_POINTS) -> int:
    """
    Locate the V1→V2 breakpoint on one wing.

    1. Grow a far-end seed until its regression slope is positive.
    2. Extend leftward while slope stays positive and ≤ seed × tolerance.
    3. Fall back to SSR split if no positive slope exists.
    """
    n = len(wing_offsets)
    if n < 2 * min_points:
        return _ssr_breakpoint(wing_offsets, wing_times_sec, min_points)

    seed_start: Optional[int]   = None
    seed_slope: Optional[float] = None

    for seed_size in range(min_points, n + 1):
        start   = n - seed_size
        seg_off = wing_offsets[start:]
        seg_t   = wing_times_sec[start:]
        if len(seg_off) < 2:
            continue
        sl, *_ = linregress(seg_off, seg_t)
        if sl > 0:
            seed_start, seed_slope = start, sl
            break

    if seed_slope is None:
        print("    ⚠  No positive V2 slope on wing – SSR fallback.")
        return _ssr_breakpoint(wing_offsets, wing_times_sec, min_points)

    best_start = seed_start
    for ext in range(seed_start - 1, min_points - 1, -1):
        sl, *_ = linregress(wing_offsets[ext:], wing_times_sec[ext:])
        if sl <= 0 or sl > seed_slope * BP_SLOPE_TOLERANCE:
            break
        best_start = ext
    return best_start


# ===========================================================================
# Per-wing refraction fit
# ===========================================================================

def _fit_wing(wing_geo: np.ndarray,
              wing_offsets: np.ndarray,
              wing_times_sec: np.ndarray,
              full_geo: np.ndarray,
              full_offsets: np.ndarray,
              full_times_sec: np.ndarray,
              side: str,
              min_points: int = MIN_SEGMENT_POINTS) -> WingResult:
    """
    Two-segment linear fit on one wing.
    Returns a parameter dict, or None if the wing is too short.
    """
    if len(wing_offsets) < 2 * min_points:
        return None

    v2_local_start    = find_breakpoint_on_wing(
        wing_offsets, wing_times_sec, min_points)
    wing_full_indices = np.array(
        [np.where(full_geo == g)[0][0] for g in wing_geo])

    v2_mask = np.zeros(len(full_geo), dtype=bool)
    v2_mask[wing_full_indices[v2_local_start:]] = True
    v1_mask = ~v2_mask

    slope2, intercept2, *_ = linregress(full_offsets[v2_mask],
                                        full_times_sec[v2_mask])
    slope1, intercept1, *_ = linregress(full_offsets[v1_mask],
                                        full_times_sec[v1_mask])

    for name, sl in (('V1', slope1), ('V2', slope2)):
        if sl <= 0:
            print(f"    ⚠  [{side}] {name} slope non-positive "
                  f"({sl:.6f}) – unreliable.")

    slope1 = abs(slope1) if slope1 <= 0 else slope1
    slope2 = abs(slope2) if slope2 <= 0 else slope2

    v1, v2 = 1.0 / slope1, 1.0 / slope2
    t_i    = intercept2
    depth  = (
        (t_i * v1 * v2) / (2.0 * np.sqrt(v2 ** 2 - v1 ** 2))
        if v2 >= V2_ROCK_MIN and (v2 ** 2 - v1 ** 2) > 0
        else float('nan')
    )

    return dict(
        side       = side,
        v1         = v1,
        v2         = v2,
        t_i_ms     = t_i * 1000.0,
        depth      = depth,
        slope1     = slope1,
        intercept1 = intercept1,
        slope2     = slope2,
        intercept2 = intercept2,
        bp_offset  = float(full_offsets[v2_mask].min()),
        bp_geo     = 0.0,
        v2_mask    = v2_mask,
    )


# ===========================================================================
# Plausibility checks
# ===========================================================================

def _check_wing(wing: WingResult, label: str, warnings: list[str]) -> None:
    """Append warning strings for out-of-range velocities."""
    if wing is None:
        return
    v1, v2 = wing['v1'], wing['v2']
    checks = [
        (not (V1_MIN <= v1 <= V1_MAX),
         f"⚠ {label} V1 = {v1:.0f} m/s outside "
         f"[{V1_MIN:.0f}–{V1_MAX:.0f}]"),
        (not (V2_MIN <= v2 <= V2_MAX),
         f"⚠ {label} V2 = {v2:.0f} m/s outside "
         f"[{V2_MIN:.0f}–{V2_MAX:.0f}]"),
        (v2 <= v1,
         f"⚠ {label} V2 ≤ V1 – refraction condition not met"),
    ]
    for condition, msg in checks:
        if condition:
            print(f"    {msg}")
            warnings.append(msg)


# ===========================================================================
# Shot analyser
# ===========================================================================

def analyse_shot(geophone_locs: np.ndarray,
                 times_ms: np.ndarray,
                 shot_pos: Optional[float] = None) -> ShotResult:
    """
    Full ITM analysis for one shot record.

    Returns dict with keys: right, left, true_shot_loc, offsets, warnings.
    """
    times_sec = times_ms / 1000.0
    offsets, true_shot_loc, shot_idx = compute_offsets(
        geophone_locs, times_ms, shot_pos=shot_pos)

    print(f"  Shot pos (header) : {shot_pos} m  |  "
          f"nearest geo : {geophone_locs[shot_idx]:.1f} m  |  "
          f"true shot : {true_shot_loc:.1f} m")

    left_idx  = np.where(geophone_locs < true_shot_loc)[0]
    right_idx = np.where(geophone_locs > true_shot_loc)[0]
    left_idx_s  = left_idx[np.argsort(offsets[left_idx])]
    right_idx_s = right_idx[np.argsort(offsets[right_idx])]

    print(f"  Wings → left : {len(left_idx)} pts  |  "
          f"right : {len(right_idx)} pts")

    warnings: list[str] = []

    def _process_wing(idx_sorted, side: str, sign: int) -> WingResult:
        if len(idx_sorted) == 0:
            print(f"  — {side.capitalize()} wing: no geophones – skipped.")
            return None
        print(f"  — {side.capitalize()} wing —")
        result = _fit_wing(
            wing_geo       = geophone_locs[idx_sorted],
            wing_offsets   = offsets[idx_sorted],
            wing_times_sec = times_sec[idx_sorted],
            full_geo       = geophone_locs,
            full_offsets   = offsets,
            full_times_sec = times_sec,
            side           = side,
        )
        if result is None:
            print("    Too few points – skipped.")
            return None
        result['bp_geo'] = true_shot_loc + sign * result['bp_offset']
        print(f"    V1={result['v1']:.1f}  V2={result['v2']:.1f} m/s  "
              f"t_i={result['t_i_ms']:.2f} ms  "
              f"z={result['depth']:.2f} m  "
              f"BP@{result['bp_geo']:.1f} m")
        _check_wing(result, side.capitalize(), warnings)
        return result

    return dict(
        right         = _process_wing(right_idx_s, 'right', +1),
        left          = _process_wing(left_idx_s,  'left',  -1),
        true_shot_loc = true_shot_loc,
        offsets       = offsets,
        warnings      = warnings,
    )


# ===========================================================================
# Rock-depth collector
# ===========================================================================

def collect_rock_points(records: list[dict],
                        elev_data: dict[int, tuple[np.ndarray, np.ndarray]]
                        ) -> dict[int, list[dict]]:
    """
    Convert ITM depths to absolute rock elevations per transect.

    Only wings with depth > 0 and v2 >= V2_ROCK_MIN are included.
    Uses the in-memory *elev_data* dict instead of on-disk files.
    """
    rock: dict[int, list[dict]] = {}

    for rec in records:
        tid = rec['transect_id']
        res = rec['res']

        if tid not in elev_data:
            continue
        elev_x, elev_z = elev_data[tid]

        for wing in filter(None, (res.get('right'), res.get('left'))):
            depth = wing.get('depth', float('nan'))
            v2    = wing.get('v2',    float('nan'))
            if np.isnan(depth) or depth <= 0 or v2 < V2_ROCK_MIN:
                continue

            x      = res['true_shot_loc']
            z_surf = interpolate_elevation(elev_x, elev_z, x)
            if z_surf is None:
                continue

            rock.setdefault(tid, []).append(dict(
                x_geo     = x,
                z_surface = z_surf,
                depth     = depth,
                z_rock    = z_surf - depth,
                side      = wing['side'],
                v2        = v2,
            ))

    return rock


# ===========================================================================
# Travel-time plot
# ===========================================================================

def _draw_traveltime_plot(ax: plt.Axes,
                          geophone_locs: np.ndarray,
                          times_ms: np.ndarray,
                          res: ShotResult,
                          title: str) -> None:
    """Travel-time vs. geophone-position plot for one shot."""
    true_shot_loc = res['true_shot_loc']
    warnings      = res.get('warnings', [])
    right         = res.get('right')
    left          = res.get('left')
    ref = right if right is not None else left
    if ref is None:
        ax.set_title(title + '  [no data]')
        return

    for wing, color in ((right, 'blue'), (left, 'darkgreen')):
        if wing is None:
            continue
        mask = wing['v2_mask']
        ax.scatter(geophone_locs[mask], times_ms[mask],
                   color=color, zorder=5, s=70, marker='D',
                   label=f"V2 {wing['side']} ({mask.sum()} pts)")

    ax.scatter(geophone_locs, times_ms,
               color='red', zorder=6, s=20, label='First-arrival picks')

    geo_full = np.linspace(geophone_locs.min(), geophone_locs.max(), 500)
    off_full = np.abs(geo_full - true_shot_loc)
    ax.plot(geo_full,
            (ref['slope1'] * off_full + ref['intercept1']) * 1000,
            color='green', linestyle='--', linewidth=1.5,
            label=f"V1 = {ref['v1']:.0f} m/s")

    for wing, color, geo_lo, geo_hi, sign in (
            (right, 'blue',      true_shot_loc,       geophone_locs.max(), +1),
            (left,  'darkgreen', geophone_locs.min(), true_shot_loc,       -1),
    ):
        if wing is None:
            continue
        geo_w = np.linspace(geo_lo, geo_hi, 300)
        ax.plot(geo_w,
                (wing['slope2'] * sign * (geo_w - true_shot_loc)
                 + wing['intercept2']) * 1000,
                color=color, linestyle='--', linewidth=1.5,
                label=f"V2 {wing['side']} = {wing['v2']:.0f} m/s")

    ax.axvline(x=true_shot_loc, color='purple', linestyle='-',
               linewidth=1.2, zorder=4,
               label=f'Shot @ {true_shot_loc:.1f} m')

    for wing, color in ((right, 'orange'), (left, 'saddlebrown')):
        if wing is None:
            continue
        ax.axvline(x=wing['bp_geo'], color=color, linestyle='--',
                   linewidth=1.2, zorder=4,
                   label=f"BP {wing['side']} @ {wing['bp_geo']:.1f} m")

    for wing, color in ((right, 'steelblue'), (left, 'teal')):
        if wing is None:
            continue
        ax.axhline(y=wing['t_i_ms'], color=color, linestyle=':',
                   linewidth=1.0, zorder=4,
                   label=f"t_i {wing['side']} = {wing['t_i_ms']:.1f} ms")
        ax.scatter([true_shot_loc], [wing['t_i_ms']],
                   color=color, zorder=7, s=60, marker='^')

    lines: list[str] = []
    for wing in filter(None, (right, left)):
        lines.append(
            f"{wing['side'].capitalize()}: "
            f"V1={wing['v1']:.0f}  V2={wing['v2']:.0f} m/s  "
            f"t_i={wing['t_i_ms']:.1f} ms  z={wing['depth']:.2f} m")
    if right is not None and left is not None:
        z_avg = np.nanmean([right['depth'], left['depth']])
        lines.append(f"z avg = {z_avg:.2f} m  |  "
                     f"Shot @ {true_shot_loc:.1f} m")

    ax.annotate('\n'.join(lines),
                xy=(0.02, 0.97), xycoords='axes fraction', fontsize=7.5,
                verticalalignment='top', family='monospace',
                bbox=dict(boxstyle='round,pad=0.4',
                          facecolor='wheat', alpha=0.9))
    if warnings:
        ax.annotate('\n'.join(warnings),
                    xy=(0.02, 0.02), xycoords='axes fraction', fontsize=7,
                    verticalalignment='bottom', color='darkred',
                    bbox=dict(boxstyle='round,pad=0.4', facecolor='#FFE0E0',
                              edgecolor='red', linewidth=1.2, alpha=0.95))

    ax.set_xlabel('Geophone position (m)')
    ax.set_ylabel('Travel time (ms)')
    ax.set_title(title, color='darkred' if warnings else 'black')
    ax.legend(fontsize=7, loc='lower right')
    ax.grid(True, alpha=0.4)


# ===========================================================================
# Elevation + geotech plot
# ===========================================================================

def _draw_geotech_overlays(ax: plt.Axes,
                            elev_x: np.ndarray,
                            elev_z: np.ndarray,
                            geotech_pts: list[GeotechPoint],
                            sheetpile: Optional[SheetPileLine]) -> None:
    """
    Overlay sheet-pile line and geotech test lines on an elevation axes.

    Rules
    -----
    - Sheet-pile: dashed vertical line from surface to SHEET_PILE_DEPTH.
    - Each test: solid vertical line from surface to tested_depth.
    - Rock found: star marker at the rock depth with depth annotation.
    - No rock found: NO point symbol at the bottom of the line.
    """
    # ── Sheet-pile ─────────────────────────────────────────────────────────────
    if sheetpile is not None:
        x_sp  = sheetpile.dist_x
        z_top = interpolate_elevation(elev_x, elev_z, x_sp)
        if z_top is not None:
            z_bot = z_top - SHEET_PILE_DEPTH
            ax.plot([x_sp, x_sp], [z_top, z_bot],
                    color=_SHEET_PILE_COLOR, linestyle='--',
                    linewidth=1.8, zorder=6,
                    label=f"Sheet pile @ x={x_sp:.1f} m "
                          f"(d={SHEET_PILE_DEPTH:.0f} m)")

    # ── Geotech tests ─────────────────────────────────────────────────────────
    legend_added: set[str] = set()

    for pt in geotech_pts:
        color  = (_DPL_COLOR if pt.test_type.upper() == 'DPL'
                  else _CPTU_COLOR)
        z_surf = interpolate_elevation(elev_x, elev_z, pt.dist_x)
        if z_surf is None:
            continue

        z_bot = z_surf - pt.tested_depth

        # Vertical test line
        line_label = (pt.test_type
                      if pt.test_type not in legend_added
                      else '_nolegend_')
        ax.plot([pt.dist_x, pt.dist_x], [z_surf, z_bot],
                color=color, linewidth=1.4, zorder=5, label=line_label)
        legend_added.add(pt.test_type)

        # Rock marker + label (only when rock was encountered)
        if pt.depth_of_rock is not None:
            z_rock      = z_surf - pt.depth_of_rock
            rock_label  = ('Rock (geotech)'
                           if 'Rock (geotech)' not in legend_added
                           else '_nolegend_')
            ax.scatter([pt.dist_x], [z_rock],
                       color=color, marker='*',
                       s=_ROCK_MARKER_SIZE ** 2,
                       zorder=7, label=rock_label)
            legend_added.add('Rock (geotech)')
            ax.annotate(f"{pt.depth_of_rock:.1f} m",
                        xy=(pt.dist_x, z_rock),
                        xytext=(4, 3), textcoords='offset points',
                        fontsize=6, color=color, clip_on=True)


def _draw_elevation_plot(ax: plt.Axes,
                         elev_x: np.ndarray,
                         elev_z: np.ndarray,
                         rock_points: list[dict],
                         title: str,
                         geotech_pts: Optional[list[GeotechPoint]] = None,
                         sheetpile: Optional[SheetPileLine] = None,
                         ) -> None:
    """
    Ground-surface profile with ITM rock-depth points and geotech overlays.

    The elevation data may not span the full plot x-range [ELEV_X_MIN,
    ELEV_X_MAX].  Where it falls short the surface line is extended
    horizontally (clamped) to the nearest known elevation so the fill and
    profile always reach both axis edges.

    Fixed axes  x: ELEV_X_MIN → ELEV_X_MAX
                y: ELEV_Y_MIN → ELEV_Y_MAX
    with 1:1 data-unit aspect ratio enforced by set_aspect('equal').
    """
    if len(elev_x) == 0:
        ax.set_title(title + '  [no elevation data]')
        return

    # Build a dense x-grid that covers the full plot range, then
    # use np.interp (which clamps at the boundary values) to get z.
    # This automatically extends the profile line to both edges.
    x_plot = np.linspace(ELEV_X_MIN, ELEV_X_MAX, 1000)
    z_plot = np.interp(x_plot, elev_x, elev_z)   # clamps outside data range

    # ── Ground surface ────────────────────────────────────────────────────────
    ax.fill_between(x_plot, z_plot, ELEV_Y_MIN,
                    color='#D2B48C', alpha=0.45, label='Ground surface')
    ax.plot(x_plot, z_plot,
            color='saddlebrown', linewidth=1.8, label='Elevation profile')

    # Mark the extent of actual measured data vs. extrapolated ends
    x_data_min, x_data_max = elev_x.min(), elev_x.max()
    for x_lo, x_hi in (
            (ELEV_X_MIN, min(x_data_min, ELEV_X_MAX)),
            (max(x_data_max, ELEV_X_MIN), ELEV_X_MAX),
    ):
        if x_lo < x_hi:
            mask = (x_plot >= x_lo) & (x_plot <= x_hi)
            ax.plot(x_plot[mask], z_plot[mask],
                    color='saddlebrown', linewidth=1.2,
                    linestyle=':', alpha=0.6)   # dotted = extrapolated

    # ── ITM rock-depth points ─────────────────────────────────────────────────
    if rock_points:
        pts    = sorted(rock_points, key=lambda p: p['x_geo'])
        x_rock = np.array([p['x_geo']  for p in pts])
        z_rock = np.array([p['z_rock'] for p in pts])

        ax.plot(x_rock, z_rock,
                color='dimgray', linestyle='--', linewidth=0.9,
                alpha=0.6, zorder=3)

        for side, color, label in (
                ('right', 'steelblue', 'ITM rock (right wing)'),
                ('left',  'darkgreen', 'ITM rock (left wing)'),
        ):
            xs = [p['x_geo']  for p in pts if p['side'] == side]
            zs = [p['z_rock'] for p in pts if p['side'] == side]
            if xs:
                ax.scatter(xs, zs, color=color, marker='D',
                           s=55, zorder=5, label=label)

        for p in pts:
            ax.annotate(f"{p['depth']:.1f} m",
                        xy=(p['x_geo'], p['z_rock']),
                        xytext=(4, -10), textcoords='offset points',
                        fontsize=6.5, color='navy', clip_on=True)

    # ── Geotech overlays ──────────────────────────────────────────────────────
    # Pass the full (possibly extrapolated) x/z arrays so that
    # interpolate_elevation works correctly for any x in [0, 28].
    _draw_geotech_overlays(ax, x_plot, z_plot, geotech_pts or [], sheetpile)

    # ── Fixed 1:1 axes ────────────────────────────────────────────────────────
    ax.set_xlim(ELEV_X_MIN, ELEV_X_MAX)
    ax.set_ylim(ELEV_Y_MIN, ELEV_Y_MAX)
    ax.set_aspect('equal', adjustable='box')

    ax.set_xlabel('Distance along transect (m)', fontsize=8)
    ax.set_ylabel('Elevation (m ASL)', fontsize=8)
    ax.set_title(title, fontsize=9)
    ax.legend(fontsize=6, loc='lower right')
    ax.grid(True, alpha=0.35)
    ax.tick_params(labelsize=7)


# ===========================================================================
# PDF page layout
# ===========================================================================

# A4 portrait in inches
_A4_W, _A4_H = 8.27, 11.69

# Travel-time page margins (figure-fraction)
_TT_L, _TT_R  = 0.10, 0.97
_TT_B, _TT_T  = 0.04, 0.97
_TT_GAP       = 0.06

# Elevation page: we must honour a 1:1 data aspect for a 28×22 m data window.
# Data aspect ratio = 28 / 22 ≈ 1.273  (wider than tall in data units).
# On A4 (8.27 in wide), usable width ≈ 7.5 in  → plot height ≈ 7.5/1.273 ≈ 5.89 in.
# With two plots + gap on 11.69 in page:  2×5.89 + gap < 11.69  ✓ (gap ≈ 0.4 in)
_ELEV_MARGIN_IN = 0.55   # inches on each side / top / bottom
_ELEV_GAP_IN    = 0.50   # inches between the two plots


def _elev_page_layout() -> tuple[plt.Figure,
                                  list[plt.Axes]]:
    """
    Create one A4 figure with two elevation axes that each have a true
    1:1 data-unit aspect ratio for the 28 m × 22 m data window.

    Returns (fig, [ax_top, ax_bottom]).
    """
    data_w = ELEV_X_MAX - ELEV_X_MIN          # 28 m
    data_h = ELEV_Y_MAX - ELEV_Y_MIN          # 22 m
    aspect = data_w / data_h                   # ≈ 1.273

    usable_w_in = _A4_W - 2 * _ELEV_MARGIN_IN
    plot_w_in   = usable_w_in
    plot_h_in   = plot_w_in / aspect           # true 1:1 size

    # Total height needed for two plots
    total_h_in  = (2 * plot_h_in
                   + _ELEV_GAP_IN
                   + 2 * _ELEV_MARGIN_IN)

    # If it doesn't fit on A4, scale down proportionally
    if total_h_in > _A4_H:
        scale       = _A4_H / total_h_in
        plot_w_in  *= scale
        plot_h_in  *= scale

    fig_h = max(_A4_H,
                2 * plot_h_in + _ELEV_GAP_IN + 2 * _ELEV_MARGIN_IN)
    fig   = plt.figure(figsize=(_A4_W, fig_h))
    fig.patch.set_facecolor('white')

    # Convert to figure fractions
    def _frac(inches: float) -> float:
        return inches / fig_h

    bot_lower = _frac(_ELEV_MARGIN_IN)
    bot_upper = _frac(_ELEV_MARGIN_IN + plot_h_in + _ELEV_GAP_IN)
    left_f    = _ELEV_MARGIN_IN / _A4_W
    w_f       = plot_w_in / _A4_W
    h_f       = _frac(plot_h_in)

    ax_bot = fig.add_axes([left_f, bot_lower, w_f, h_f])
    ax_top = fig.add_axes([left_f, bot_upper, w_f, h_f])

    return fig, [ax_top, ax_bot]


# ===========================================================================
# PDF writers
# ===========================================================================

def save_traveltime_pdf(records: list[dict], pdf_path: Path) -> None:
    """Write one travel-time plot per shot, two per A4 page."""
    plot_w = _TT_R - _TT_L
    plot_h = (_TT_T - _TT_B - _TT_GAP) / 2.0
    bottoms = [_TT_B + plot_h + _TT_GAP, _TT_B]

    with pdf_backend.PdfPages(pdf_path) as pdf:
        prev_tid = None
        i = 0
        while i < len(records):
            tid = records[i]['transect_id']

            if tid != prev_tid:
                sep = plt.figure(figsize=(_A4_W, 0.8))
                sep.text(0.5, 0.5, f"Transect  {tid}",
                         ha='center', va='center',
                         fontsize=14, fontweight='bold',
                         transform=sep.transFigure)
                pdf.savefig(sep, bbox_inches='tight')
                plt.close(sep)
                prev_tid = tid

            fig = plt.figure(figsize=(_A4_W, _A4_H))
            fig.patch.set_facecolor('white')

            for slot in range(2):
                if i >= len(records) or records[i]['transect_id'] != tid:
                    break
                r  = records[i]
                ax = fig.add_axes([_TT_L, bottoms[slot], plot_w, plot_h])
                _draw_traveltime_plot(
                    ax, r['geophone_locs'], r['times_ms'], r['res'],
                    title=f"Refraction ITM – {r['file_name']}")
                i += 1

            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)

    print(f"  Travel-time PDF saved → {pdf_path}")


def save_elevation_pdf(records: list[dict],
                       elev_data: dict[int, tuple[np.ndarray, np.ndarray]],
                       pdf_path: Path,
                       geotech_by_tid: Optional[
                           dict[int, list[GeotechPoint]]] = None,
                       sheetpile_by_tid: Optional[
                           dict[int, SheetPileLine]] = None,
                       ) -> None:
    """
    Write one elevation + rock-depth profile per transect, two per A4 page.
    Each plot has a true 1:1 data-unit aspect ratio.
    """
    rock_by_tid   = collect_rock_points(records, elev_data)
    transect_ids  = list(dict.fromkeys(r['transect_id'] for r in records))

    # Also include transects that have geotech data but no seismic records
    if geotech_by_tid:
        for tid in geotech_by_tid:
            if tid not in transect_ids:
                transect_ids.append(tid)

    if not transect_ids:
        print("  No transect data – skipping elevation PDF.")
        return

    with pdf_backend.PdfPages(pdf_path) as pdf:
        i = 0
        while i < len(transect_ids):
            fig, axes = _elev_page_layout()

            for ax in axes:
                if i >= len(transect_ids):
                    ax.set_visible(False)
                    continue

                tid = transect_ids[i]
                i  += 1

                elev_x, elev_z = elev_data.get(tid,
                                               (np.array([]), np.array([])))
                if len(elev_x) == 0:
                    print(f"  ⚠  No elevation data for transect {tid}")

                _draw_elevation_plot(
                    ax, elev_x, elev_z,
                    rock_by_tid.get(tid, []),
                    title=f"Transect {tid} — Ground profile & rock depth",
                    geotech_pts = (geotech_by_tid or {}).get(tid),
                    sheetpile   = (sheetpile_by_tid or {}).get(tid),
                )

            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)

    print(f"  Elevation PDF saved → {pdf_path}")


# ===========================================================================
# Excel writer
# ===========================================================================

_HEADERS = [
    "Transect", "File", "Shot pos (m)",
    "V1 right (m/s)", "V2 right (m/s)", "t_i right (ms)", "Depth right (m)",
    "V1 left (m/s)",  "V2 left (m/s)",  "t_i left (ms)",  "Depth left (m)",
    "Depth avg (m)", "Warnings",
]
_COL_WIDTHS = [10, 24, 13, 14, 14, 14, 14, 13, 13, 13, 13, 13, 50]


def _wing_val(wing: WingResult, key: str, decimals: int = 2):
    """Safe getter – returns 'N/A' for missing / NaN values."""
    if wing is None:
        return 'N/A'
    val = wing.get(key, float('nan'))
    if isinstance(val, float):
        return round(val, decimals) if not np.isnan(val) else 'N/A'
    return val


def save_excel(results: list[dict], xlsx_path: Path) -> None:
    """Write a formatted summary spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Refraction ITM Results"

    thin      = Side(style='thin')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre    = Alignment(horizontal="center")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")

    for col, (header, width) in enumerate(
            zip(_HEADERS, _COL_WIDTHS), start=1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30

    for row_num, r in enumerate(results, start=2):
        res   = r['res']
        right = res.get('right')
        left  = res.get('left')
        depths = [w['depth'] for w in filter(None, (right, left))]
        d_avg  = float(np.nanmean(depths)) if depths else float('nan')

        row_values = [
            r['transect_id'], r['file_name'],
            round(res['true_shot_loc'], 3),
            _wing_val(right, 'v1', 1), _wing_val(right, 'v2', 1),
            _wing_val(right, 't_i_ms', 3), _wing_val(right, 'depth', 3),
            _wing_val(left,  'v1', 1), _wing_val(left,  'v2', 1),
            _wing_val(left,  't_i_ms', 3), _wing_val(left,  'depth', 3),
            round(d_avg, 3) if not np.isnan(d_avg) else 'N/A',
            ' | '.join(res.get('warnings', [])),
        ]
        has_warnings = bool(res.get('warnings'))
        for col, value in enumerate(row_values, start=1):
            cell           = ws.cell(row=row_num, column=col, value=value)
            cell.border    = border
            cell.alignment = centre
            if has_warnings:
                cell.fill = warn_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"
    wb.save(xlsx_path)
    print(f"  Excel saved → {xlsx_path}")


# ===========================================================================
# Main
# ===========================================================================

def main() -> int:
    script_dir = Path(__file__).parent
    pick_dir   = script_dir / 'pick_files'
    work_dir   = pick_dir if pick_dir.is_dir() else script_dir

    vs_files = sorted(work_dir.glob('*.vs'))
    if not vs_files:
        print("No .vs files found.")
        return 1

    print(f"Found {len(vs_files)} .vs file(s).\n")
    transects = group_by_transect(vs_files)
    print(f"Transects found: {list(transects.keys())}\n")

    # ── Elevation data from elev.csv ──────────────────────────────────────────
    elev_data = read_elev_csv(script_dir / 'elev.csv')

    # ── Geotech data from geotech.csv (optional) ──────────────────────────────
    geotech_by_tid, sheetpile_by_tid = read_geotech_csv(
        script_dir / 'geotech.csv')

    # ── Process all shots ─────────────────────────────────────────────────────
    all_records: list[dict] = []

    for tid, files in transects.items():
        print(f"\n{'═' * 60}")
        print(f"  TRANSECT  {tid}  ({len(files)} shots)")
        print(f"{'═' * 60}\n")

        for path in sorted(files):
            print(f"  {'─' * 50}")
            print(f"  Processing : {path.name}")

            try:
                shot_pos, distances, times_ms = read_vs_file(path)
            except Exception as exc:
                print(f"  ✗ Read error: {exc}")
                continue

            if times_ms.size == 0:
                print("  ✗ No data – skipping.")
                continue

            try:
                res = analyse_shot(distances, times_ms, shot_pos=shot_pos)
            except Exception as exc:
                print(f"  ✗ Analysis error: {exc}")
                continue

            all_records.append(dict(
                transect_id   = tid,
                file_name     = path.name,
                geophone_locs = distances,
                times_ms      = times_ms,
                res           = res,
            ))

    if not all_records:
        print("\nNo files processed successfully.")
        return 1

    print(f"\n{'═' * 60}")
    print(f"Processed {len(all_records)} / {len(vs_files)} shots.\n")

    save_traveltime_pdf(all_records,
                        script_dir / "refra_itm_traveltimes.pdf")
    save_excel(all_records,
               script_dir / "refra_itm_results.xlsx")
    save_elevation_pdf(
        all_records,
        elev_data,
        script_dir / "refra_itm_elevation.pdf",
        geotech_by_tid   = geotech_by_tid   or None,
        sheetpile_by_tid = sheetpile_by_tid  or None,
    )

    return 0


if __name__ == "__main__":
    sys.exit(main())